import openpyxl
import matplotlib.pyplot as plt
import matplotlib

matplotlib.use('TkAgg')


def load_waveform_from_sheet(ws):
    voltages = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header
        _, voltage, _ = row
        voltages.append(voltage)
    return voltages

def insert_data_points(num_points=10, value=258000000):
    return [value] * num_points

def plot_waveforms(input_filename):
    wb = openpyxl.load_workbook(input_filename)
    sheetnames = wb.sheetnames

    all_voltages = []
    total_points = 0

    for sheetname in sheetnames:
        ws = wb[sheetname]
        voltages = load_waveform_from_sheet(ws)
        if all_voltages:
            # Insert data points between waveforms
            insert_voltages = insert_data_points()
            all_voltages.extend(insert_voltages)
            total_points += len(insert_voltages)

        all_voltages.extend(voltages)
        total_points += len(voltages)

    # Create a new x-axis based on the total number of points
    all_times = list(range(total_points))

    plt.figure(figsize=(14, 7))
    plt.plot(all_times, all_voltages, label='Combined Waveforms')

    plt.xlabel('Index')
    plt.ylabel('Voltage')
    plt.title('Combined Waveforms')
    plt.legend()
    plt.show()


input_filename = r'C:\Users\dell\Desktop\training_data\training_data_1.xlsx'
plot_waveforms(input_filename)