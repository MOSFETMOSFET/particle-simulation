import random
import openpyxl


def generate_waveform_data(num_points):
    base_voltage = random.uniform(258000000, 259000000)
    dip_depth_percentage = random.uniform(0.05, 0.2)
    dip_min_voltage = base_voltage * (1 - dip_depth_percentage)

    dip_start = int(num_points * 0.001)  # start of the dip (10% of the points)
    dip_end = int(num_points * 0.999)  # end of the dip (90% of the points)

    times = [5.4 + i * 0.05 for i in range(num_points)]
    voltages = []

    for i in range(num_points):
        if dip_start <= i <= dip_end:
            progress = (i - dip_start) / (dip_end - dip_start) * 2 - 1
            voltage = base_voltage - (base_voltage - dip_min_voltage) * (1 - progress ** 2)
        else:
            voltage = base_voltage
        voltages.append(voltage)

    labels = [0] * num_points
    return times, voltages, labels


def save_waveforms_to_excel(num_waveforms, min_points, max_points, filename):
    wb = openpyxl.Workbook()
    for i in range(num_waveforms):
        ws = wb.create_sheet(title=f'Waveform_{i + 1}')
        num_points = random.randint(min_points, max_points)
        times, voltages, labels = generate_waveform_data(num_points)
        ws.append(["Time", "Voltage", "Label"])
        for time, voltage, label in zip(times, voltages, labels):
            ws.append([time, voltage, label])

    # Remove the default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save(filename)


# Number of waveforms to generate
num_waveforms = 500

# Minimum and maximum number of points in each waveform
min_points = 10
max_points = 20

# Output filename
filename = r'C:\Users\dell\Desktop\training_data\training_data_1.xlsx'

save_waveforms_to_excel(num_waveforms, min_points, max_points, filename)

print(f'Generated {num_waveforms} waveforms with points in range ({min_points}, {max_points}) and saved to {filename}')