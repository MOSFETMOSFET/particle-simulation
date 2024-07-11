import numpy as np
import matplotlib.pyplot as plt
import matplotlib
import openpyxl
import random

matplotlib.use('TkAgg')

def generate_waveform(t):
    """
    Generates a waveform with two local minima and highest values at the edges, without random noise.
    """
    # Create a base waveform with two dips
    center1 = 2 + np.random.uniform(-0.5, 0.5)
    center2 = 4 + np.random.uniform(-0.5, 0.5)
    dip_width1 = 0.5 + np.random.uniform(-0.2, 0.2)
    dip_width2 = 0.5 + np.random.uniform(-0.2, 0.2)
    amplitude1 = 1.6e8 + np.random.uniform(-0.25e8, 0.25e8)
    amplitude2 = 1.6e8 + np.random.uniform(-0.25e8, 0.25e8)

    waveform = 2.59e8 - amplitude1 * np.exp(-((t - center1) ** 2) / (2 * dip_width1 ** 2)) - amplitude2 * np.exp(
        -((t - center2) ** 2) / (2 * dip_width2 ** 2))

    return waveform

def save_waveforms_to_excel(t, num_waveforms, filename):
    """
    Saves the waveforms data to an Excel file, each waveform in a separate sheet.
    """
    wb = openpyxl.Workbook()
    for i in range(num_waveforms):
        ws = wb.create_sheet(title=f'Waveform_{i + 1}')
        waveform = generate_waveform(t)
        ws.append(["Time", "Voltages", "Label"])
        for time, value in zip(t, waveform):
            ws.append([time, value, 1])

    # Remove the default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save(filename)

def main():
    # Time array
    t = np.linspace(0, 6, 1000)

    # Number of waveforms to generate
    num_waveforms = 10

    excel_filename = r'C:\Users\dell\Desktop\training_data\test_data.xlsx'
    #excel_filename = r'C:\Users\dell\Desktop\training_data\training_data_2.xlsx'

    save_waveforms_to_excel(t, num_waveforms, excel_filename)

if __name__ == "__main__":
    main()
